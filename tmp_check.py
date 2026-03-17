# -*- coding: utf-8 -*-
import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

chart_xlsx = r'H:\マイドライブ\〇市道 南千反畑町第１号線舗装補修工事\１６出来形管理\出来形管理図表、舗装工、南千反畑.xlsx'
soukatu_xlsx = r'H:\マイドライブ\〇市道 南千反畑町第１号線舗装補修工事\１６出来形管理\出来形管理総括表.xlsx'

wb_chart = openpyxl.load_workbook(chart_xlsx, data_only=True)
wb_soukatu = openpyxl.load_workbook(soukatu_xlsx, data_only=False)
ws_s = wb_soukatu['総括表2']

def compare(label, computed, soukatu_val):
    if soukatu_val is None:
        return f"✗(None→{computed})"
    if isinstance(computed, float):
        if abs(computed - float(soukatu_val)) < 0.01:
            return "✓"
    elif computed == soukatu_val:
        return "✓"
    return f"✗({soukatu_val}→{computed})"

# --- 切削 基準高V (V1-V5) ---
ws = wb_chart['切削基準高V']
print('=== 切削 基準高V ===')
for vi, (des_r, act_r) in enumerate([(7,8),(10,11),(13,14),(16,17),(19,20)], 1):
    diffs = []
    for col in range(6, 13):
        d = ws.cell(row=des_r, column=col).value
        a = ws.cell(row=act_r, column=col).value
        if d is not None and a is not None:
            diffs.append(round((a - d) * 1000))
    n = len(diffs)
    if n == 0:
        continue
    avg = round(sum(diffs)/n, 1)
    mx = max(diffs)
    mn = min(diffs)
    b80 = sum(1 for d in diffs if -24 <= d <= 24)
    c50 = sum(1 for d in diffs if -15 <= d <= 15)

    s_row = 9 + (vi-1)*2
    ck = compare('avg', avg, ws_s[f'K{s_row}'].value)
    cl = compare('max', mx, ws_s[f'L{s_row}'].value)
    cm = compare('min', mn, ws_s[f'M{s_row}'].value)
    cn = compare('N', n, ws_s[f'N{s_row}'].value)
    co = compare('B80', b80, ws_s[f'O{s_row}'].value)
    cq = compare('C50', c50, ws_s[f'Q{s_row}'].value)
    print(f'  V{vi}(Row{s_row}): avg{ck} max{cl} min{cm} N{cn} B80{co} C50{cq}')

# --- 切削 幅員W (W1-W2) ---
ws = wb_chart['切削幅員W']
print('\n=== 切削 幅員W ===')
for wi, diff_r in enumerate([9, 12], 1):
    diffs = []
    for col in range(6, 13):
        v = ws.cell(row=diff_r, column=col).value
        if v is not None and isinstance(v, (int, float)):
            diffs.append(v)
    n = len(diffs)
    if n == 0:
        continue
    avg = round(sum(diffs)/n, 1)
    mx = max(diffs)
    mn = min(diffs)
    b80 = sum(1 for d in diffs if d >= -16)
    c50 = sum(1 for d in diffs if d >= -10)

    s_row = 19 + (wi-1)*2
    ck = compare('avg', avg, ws_s[f'K{s_row}'].value)
    cl = compare('max', mx, ws_s[f'L{s_row}'].value)
    cm = compare('min', mn, ws_s[f'M{s_row}'].value)
    cn = compare('N', n, ws_s[f'N{s_row}'].value)
    co = compare('B80', b80, ws_s[f'O{s_row}'].value)
    cq = compare('C50', c50, ws_s[f'Q{s_row}'].value)
    print(f'  W{wi}(Row{s_row}): avg{ck} max{cl} min{cm} N{cn} B80{co} C50{cq}')

# --- 表層 基準高V (V1-V5) ---
ws = wb_chart['表層V']
print('\n=== 表層 基準高V ===')
for vi, (des_r, act_r) in enumerate([(7,8),(10,11),(13,14),(16,17),(19,20)], 1):
    diffs = []
    for col in range(6, 13):
        d = ws.cell(row=des_r, column=col).value
        a = ws.cell(row=act_r, column=col).value
        if d is not None and a is not None:
            diffs.append(round((a - d) * 1000))
    n = len(diffs)
    if n == 0:
        continue
    avg = round(sum(diffs)/n, 1)
    mx = max(diffs)
    mn = min(diffs)
    b80 = sum(1 for d in diffs if -24 <= d <= 24)
    c50 = sum(1 for d in diffs if -15 <= d <= 15)

    s_row = 23 + (vi-1)*2
    ck = compare('avg', avg, ws_s[f'K{s_row}'].value)
    cl = compare('max', mx, ws_s[f'L{s_row}'].value)
    cm = compare('min', mn, ws_s[f'M{s_row}'].value)
    cn = compare('N', n, ws_s[f'N{s_row}'].value)
    co = compare('B80', b80, ws_s[f'O{s_row}'].value)
    cq = compare('C50', c50, ws_s[f'Q{s_row}'].value)
    print(f'  V{vi}(Row{s_row}): avg{ck} max{cl} min{cm} N{cn} B80{co} C50{cq}')

# --- 表層 幅員W (W1-W2) ---
ws = wb_chart['表層W']
print('\n=== 表層 幅員W ===')
for wi, diff_r in enumerate([9, 12], 1):
    diffs = []
    for col in range(6, 13):
        v = ws.cell(row=diff_r, column=col).value
        if v is not None and isinstance(v, (int, float)):
            diffs.append(v)
    n = len(diffs)
    if n == 0:
        continue
    avg = round(sum(diffs)/n, 1)
    mx = max(diffs)
    mn = min(diffs)

    s_row = 33 + (wi-1)*2
    ck = compare('avg', avg, ws_s[f'K{s_row}'].value)
    cl = compare('max', mx, ws_s[f'L{s_row}'].value)
    cm = compare('min', mn, ws_s[f'M{s_row}'].value)
    cn = compare('N', n, ws_s[f'N{s_row}'].value)
    print(f'  W{wi}(Row{s_row}): avg{ck} max{cl} min{cm} N{cn}')

# --- 厚さtコア (Row37) ---
ws = wb_chart['厚さtコア']
print('\n=== 厚さtコア (Row37) ===')
diffs = []
for col in range(6, 9):
    v = ws.cell(row=9, column=col).value
    if v is not None and isinstance(v, (int, float)):
        diffs.append(v)
n = len(diffs)
print(f'  data={diffs}')
if n > 0:
    avg = round(sum(diffs)/n, 1)
    mx = max(diffs)
    mn = min(diffs)
    b80 = sum(1 for d in diffs if d >= -5.6)
    c50 = sum(1 for d in diffs if d >= -3.5)
    print(f'  computed: avg={avg} max={mx} min={mn} n={n} B80={b80} C50={c50}')
    s = ws_s
    print(f'  総括表:   avg={s["K37"].value} max={s["L37"].value} min={s["M37"].value} N={s["N37"].value} O={s["O37"].value} Q={s["Q37"].value}')
    ck = compare('avg', avg, s['K37'].value)
    cl = compare('max', mx, s['L37'].value)
    cm = compare('min', mn, s['M37'].value)
    cn = compare('N', n, s['N37'].value)
    print(f'  result: avg{ck} max{cl} min{cm} N{cn}')

# --- 延長 (Row39) ---
print('\n=== 延長 (Row39) ===')
ws_name = '延長L' if '延長L' in wb_chart.sheetnames else '延長'
ws = wb_chart[ws_name]
print(f'  sheet: {ws_name}')
for r in range(6, 13):
    vals = []
    for c in range(5, 13):
        v = ws.cell(row=r, column=c).value
        if v is not None:
            cl = chr(64+c)
            vals.append(f'{cl}{r}={v}')
    if vals:
        print(f'  {", ".join(vals)}')
print(f'  総括表: avg={ws_s["K39"].value} max={ws_s["L39"].value} min={ws_s["M39"].value} N={ws_s["N39"].value}')

# --- 平坦性σ (Row41) ---
ws = wb_chart['平坦性σ']
print('\n=== 平坦性σ (Row41) ===')
diffs = []
for col in range(6, 8):
    v = ws.cell(row=9, column=col).value
    if v is not None and isinstance(v, (int, float)):
        diffs.append(v)
print(f'  data={diffs}')
if diffs:
    avg = round(sum(diffs)/len(diffs), 2)
    mx = max(diffs)
    mn = min(diffs)
    n = len(diffs)
    shanai = sum(1 for d in diffs if d <= 1.9)
    c50_val = sum(1 for d in diffs if d <= 1.2)
    print(f'  computed: avg={avg} max={mx} min={mn} n={n} shanai={shanai} C50={c50_val}')
    print(f'  総括表:   avg={ws_s["K41"].value} max={ws_s["L41"].value} min={ws_s["M41"].value} N={ws_s["N41"].value} O={ws_s["O41"].value} Q={ws_s["Q41"].value}')

# --- 区画線 (Row43,45) ---
ws = wb_chart['区画線']
print('\n=== 区画線 ===')
t_diffs = []
w_diffs = []
for col in range(6, 9):
    v = ws.cell(row=9, column=col).value
    if v is not None and isinstance(v, (int, float)):
        w_diffs.append(v)
    v = ws.cell(row=12, column=col).value
    if v is not None and isinstance(v, (int, float)):
        t_diffs.append(v)
print(f'  W差={w_diffs}, t差={t_diffs}')
for label, diffs, s_row in [('厚さt', t_diffs, 43), ('幅W', w_diffs, 45)]:
    n = len(diffs)
    if n == 0:
        continue
    avg = round(sum(diffs)/n, 2)
    mx = max(diffs)
    mn = min(diffs)
    ok = sum(1 for d in diffs if d >= 0)
    print(f'  {label}(Row{s_row}): computed avg={avg} max={mx} min={mn} n={n} ok={ok}')
    print(f'  {label}(Row{s_row}): 総括表  avg={ws_s[f"K{s_row}"].value} max={ws_s[f"L{s_row}"].value} min={ws_s[f"M{s_row}"].value} N={ws_s[f"N{s_row}"].value} O={ws_s[f"O{s_row}"].value}')
